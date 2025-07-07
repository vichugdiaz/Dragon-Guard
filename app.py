# ============================================
# app.py - Sistema Dragon Guard (versión corregida)
# ============================================

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file, flash, Response
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from sqlalchemy import inspect
from datetime import datetime, timedelta, date
from io import BytesIO
from fpdf import FPDF
import os, json, pandas as pd, sqlite3
from openpyxl import Workbook
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
import shutil


# =====================
# CONFIGURACIÓN BÁSICA
# =====================

app = Flask(__name__)
app.secret_key = 'dragon_guard_secret'
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'instance', 'dragon_guard.db')
db = SQLAlchemy(app)

RUTA_LOGOS = os.path.join("static", "logos")
CONFIG_PATH = os.path.join(basedir, "config.json")
AUTO_CONFIG_PATH = os.path.join(basedir, "auto_config.json")
scheduler = BackgroundScheduler()
scheduler.start()

# =====================
# CONFIGURACIÓN VISUAL Y LICENCIA
# =====================

CONFIG_PREDETERMINADA = {
    "nombre_empresa": "Dragon Guard",
    "logo_filename": "default_logo.png",
    "estilos": {
        "color_primario": "#0d6efd",
        "color_secundario": "#6c757d",
        "fuente": "Arial, sans-serif"
    },
    "licencia": ""
}

def cargar_config():
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)
    except:
        config = CONFIG_PREDETERMINADA

    for key in CONFIG_PREDETERMINADA:
        if key not in config:
            config[key] = CONFIG_PREDETERMINADA[key]

    return config

def guardar_config(config):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4, ensure_ascii=False)

def dias_restantes_licencia():
    config = cargar_config()
    try:
        fecha_fin = datetime.strptime(config.get("licencia", ""), "%Y-%m-%d").date()
        hoy = datetime.now().date()
        return max((fecha_fin - hoy).days, 0)
    except:
        return None

# =====================
# MODELOS DE BASE DE DATOS
# =====================

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario = db.Column(db.String(50), nullable=False)
    contrasena = db.Column(db.String(50), nullable=False)
    rol = db.Column(db.String(20), nullable=False)

class Registro(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cedula = db.Column(db.String(20), nullable=False)
    empleado = db.Column(db.String(100), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    hora = db.Column(db.Time, nullable=False)
    tipo = db.Column(db.String(10), nullable=False)

class Empleado(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cedula = db.Column(db.String(20), unique=True, nullable=False)
    nombre = db.Column(db.String(100), nullable=False)

# =====================
# INICIALIZACIÓN DE BASE DE DATOS
# =====================

def tabla_existe(nombre):
    inspector = inspect(db.engine)
    return nombre in inspector.get_table_names()

def crear_bd_si_no_existe():
    if not os.path.exists(os.path.join(basedir, "instance")):
        os.makedirs(os.path.join(basedir, "instance"))
    with app.app_context():
        if not os.path.exists(os.path.join(basedir, "instance", "dragon_guard.db")) or not tabla_existe("empleado"):
            db.create_all()
            if not Usuario.query.filter_by(usuario="admin").first():
                db.session.add(Usuario(usuario="admin", contrasena="Admin123", rol="admin"))
                db.session.commit()

# =====================
# INFORME AUTOMÁTICO - INICIALIZACIÓN
# =====================


def generar_excel_automatico():
    # Obtener fecha actual
    fecha = datetime.now().strftime("%Y-%m-%d")
    nombre_archivo = f"reporte_diario_{fecha}.xlsx"

    # Ruta segura para guardar el archivo
    basedir = os.path.abspath(os.path.dirname(__file__))
    carpeta_destino = os.path.join(basedir, "instance")
    ruta_completa = os.path.join(carpeta_destino, nombre_archivo)

    # Crear carpeta instance si no existe
    os.makedirs(carpeta_destino, exist_ok=True)

    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe Diario"

    # Encabezado de ejemplo
    ws.append(["Cédula", "Nombre", "Hora Entrada", "Hora Salida", "Duración"])

    # Puedes reemplazar esta parte con tus registros reales
    datos = [
        ["12345678", "Juan Pérez", "08:00", "17:00", "9h"],
        ["87654321", "Ana Gómez", "08:15", "17:10", "8h 55m"]
    ]

    for fila in datos:
        ws.append(fila)

    # Guardar archivo
    wb.save(ruta_completa)
    print(f"✅ Informe automático guardado en: {ruta_completa}")


@app.route("/generar_informe_diario")
def generar_informe_diario():
    hoy = date.today().strftime("%Y-%m-%d")
    return redirect(url_for("exportar_excel", fecha_inicio=hoy, fecha_fin=hoy))


def generar_informe_programado():
    with app.app_context():
        try:
            generar_excel_automatico()
        except Exception as e:
            print(f"❌ Error al generar informe automático: {e}")


def inicializar_auto_informe():
    if os.path.exists(AUTO_CONFIG_PATH):
        with open(AUTO_CONFIG_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
            activo = data.get("activo", False)
            hora = data.get("hora", "00:00")

            if activo and hora:
                hora_dt = datetime.strptime(hora, "%H:%M").time()
                if scheduler.get_job("informe_diario"):
                    scheduler.remove_job("informe_diario")
                scheduler.add_job(
                    func=generar_informe_programado,
                    trigger="cron",
                    hour=hora_dt.hour,
                    minute=hora_dt.minute,
                    id="informe_diario",
                    replace_existing=True
                )


# =====================
# RUTAS FLASK (CONTINUACIÓN)
# =====================




@app.route("/", methods=["GET", "POST"])
def index():
    mensaje = error = None
    visual_config = cargar_config()

    # Validación de licencia
    dias = dias_restantes_licencia()
    if dias is not None and dias <= 0:
        return redirect(url_for("licencia"))

    if request.method == "POST":
        cedula = request.form.get("cedula", "").strip()
        empleado = Empleado.query.filter_by(cedula=cedula).first()

        if not empleado:
            error = "⚠️ FAVOR NOTIFICAR AL ADMINISTRADOR DEL SISTEMA PARA CREAR AL EMPLEADO"
        else:
            ahora = datetime.now()
            hoy = ahora.date()

            # Último registro del día para este empleado
            ultimo = Registro.query.filter_by(cedula=cedula, fecha=hoy).order_by(Registro.hora.desc()).first()

            tipo = "entrada" if not ultimo or ultimo.tipo == "salida" else "salida"

            nuevo_registro = Registro(
                cedula=cedula,
                empleado=empleado.nombre,
                fecha=hoy,
                hora=ahora.time(),
                tipo=tipo
            )
            db.session.add(nuevo_registro)
            db.session.commit()
            mensaje = f"✅ Registro exitoso de {empleado.nombre} - {tipo.upper()}"

    return render_template("index.html", config=visual_config, mensaje=mensaje, error=error, dias_restantes=dias)

@app.route("/logout_admin")
def logout_admin():
    session.clear()
    flash("Sesión cerrada correctamente", "success")
    return redirect(url_for("index"))

@app.route("/licencia", methods=["GET", "POST"])
def licencia():
    config = cargar_config()
    mensaje = error = None

    if request.method == "POST":
        clave = request.form.get("clave")
        dias = request.form.get("dias")

        if clave == "2185":
            try:
                dias = int(dias)
                fecha_fin = datetime.now().date() + timedelta(days=dias)
                config["licencia"] = fecha_fin.strftime("%Y-%m-%d")
                guardar_config(config)
                mensaje = f"✔️ Licencia activada hasta {fecha_fin.strftime('%d/%m/%Y')}."
            except:
                error = "⚠️ Número de días inválido."
        else:
            error = "❌ Clave de activación incorrecta."

    dias_restantes = dias_restantes_licencia()
    return render_template("license.html", config=config, mensaje=mensaje, error=error, dias_restantes=dias_restantes)



@app.route("/admin", methods=["GET", "POST"])
def login_admin():
    config = cargar_config()
    ruta_datos = os.path.join(basedir, "datos.json")

    clave_guardada = "Admin123"
    if os.path.exists(ruta_datos):
        with open(ruta_datos, "r", encoding="utf-8") as f:
            data = json.load(f)
            clave_guardada = data.get("admin_password", "Admin123")

    if request.method == "POST":
        clave = request.form.get("clave")
        if clave == clave_guardada:
            session["rol"] = "admin"
            session["usuario"] = "admin"
            return redirect(url_for("admin_panel"))
        else:
            return render_template("login_admin.html", error="Clave incorrecta", config=config)

    return render_template("login_admin.html", config=config)

@app.route("/admin_panel")
def admin_panel():
    if session.get("rol") != "admin":
        return redirect(url_for("login_admin"))

    empleados = Empleado.query.order_by(Empleado.nombre).all()
    config = cargar_config()
    dias_restantes = dias_restantes_licencia()

    auto_config = {"activo": False, "hora": ""}
    if os.path.exists(AUTO_CONFIG_PATH):
        with open(AUTO_CONFIG_PATH, "r", encoding="utf-8") as f:
            auto_config = json.load(f)

    return render_template("admin_panel.html",
                           empleados=empleados,
                           config=config,
                           dias_restantes=dias_restantes,
                           auto_config=auto_config)

@app.route("/admin_configuracion", methods=["GET", "POST"])
def admin_configuracion():
    if 'usuario' not in session or session.get("rol") != "admin":
        return redirect(url_for("index"))

    visual_config = cargar_config()

    if request.method == "POST":
        nuevo_nombre = request.form.get("nombre_empresa", "").strip()
        color_primario = request.form.get("color_primario", "#0d6efd")
        color_secundario = request.form.get("color_secundario", "#6c757d")
        fuente = request.form.get("fuente", "Arial, sans-serif")

        if nuevo_nombre:
            visual_config["nombre_empresa"] = nuevo_nombre

        visual_config["estilos"]["color_primario"] = color_primario
        visual_config["estilos"]["color_secundario"] = color_secundario
        visual_config["estilos"]["fuente"] = fuente

        if "logo" in request.files:
            logo = request.files["logo"]
            if logo.filename:
                nombre_archivo = secure_filename(logo.filename)
                ruta_logo = os.path.join("logos", nombre_archivo)
                ruta_completa = os.path.join(basedir, "static", ruta_logo)
                os.makedirs(os.path.dirname(ruta_completa), exist_ok=True)
                logo.save(ruta_completa)
                visual_config["logo_filename"] = nombre_archivo

        guardar_config(visual_config)
        session["mensaje"] = "✅ Configuración actualizada exitosamente."
        return redirect(url_for("admin_panel"))

    return render_template("config.html", config=visual_config)

@app.route("/restaurar_config")
def restaurar_config():
    config_default = {
        "nombre_empresa": "Dragon Guard",
        "logo_filename": "default_logo.png",
        "estilos": {
            "fuente": "Arial, sans-serif",
            "color_primario": "#007bff",
            "color_secundario": "#6c757d"
        },
        "licencia": ""
    }
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config_default, f, indent=4)
    flash("✔ La configuración ha sido restaurada a los valores por defecto correctamente.")
    return redirect(url_for("admin_panel"))

@app.route("/cambiar_clave_admin", methods=["POST"])
def cambiar_clave_admin():
    nueva_clave = request.form.get("nueva_clave")

    if nueva_clave:
        ruta_datos = os.path.join(basedir, "datos.json")
        data = {}
        if os.path.exists(ruta_datos):
            with open(ruta_datos, "r", encoding="utf-8") as f:
                data = json.load(f)

        data["admin_password"] = nueva_clave

        with open(ruta_datos, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

        flash("✔ Clave del administrador actualizada correctamente.", "success")
    else:
        flash("⚠ Por favor ingrese una nueva clave.", "warning")

    return redirect(url_for("admin_panel"))



@app.route("/importar_empleados", methods=["POST"])
def importar_empleados():
    if "archivo" not in request.files:
        flash("⚠ No se ha seleccionado ningún archivo.", "warning")
        return redirect(url_for("admin_panel"))

    archivo = request.files["archivo"]
    if archivo.filename == "":
        flash("⚠ Nombre de archivo vacío.", "warning")
        return redirect(url_for("admin_panel"))

    if not archivo.filename.endswith(".xlsx"):
        flash("⚠ Formato no válido. Debe ser un archivo .xlsx", "warning")
        return redirect(url_for("admin_panel"))

    try:
        df = pd.read_excel(archivo)

        columnas = {col.lower().strip(): col for col in df.columns}
        cedula_col = columnas.get("cédula") or columnas.get("cedula")
        nombre_col = columnas.get("nombre")

        if not cedula_col or not nombre_col:
            flash("⚠ El archivo debe contener las columnas 'Cédula' y 'Nombre'.", "warning")
            return redirect(url_for("admin_panel"))

        importados = 0
        for _, fila in df.iterrows():
            cedula = str(fila[cedula_col]).strip()
            nombre = str(fila[nombre_col]).strip().upper()

            if cedula and nombre:
                existente = Empleado.query.filter_by(cedula=cedula).first()
                if not existente:
                    nuevo = Empleado(cedula=cedula, nombre=nombre)
                    db.session.add(nuevo)
                    importados += 1

        db.session.commit()
        flash(f"✔ {importados} empleados importados correctamente.", "success")

    except Exception as e:
        flash(f"❌ Error al importar empleados: {str(e)}", "danger")

    return redirect(url_for("admin_panel"))

@app.route("/exportar_empleados")
def exportar_empleados():
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"
    ws.append(["Cédula", "Nombre"])

    empleados = Empleado.query.order_by(Empleado.nombre.asc()).all()
    for emp in empleados:
        ws.append([emp.cedula, emp.nombre])

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    return send_file(
        archivo,
        download_name="empleados.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/editar_empleado/<int:id>", methods=["GET", "POST"])
def editar_empleado(id):
    empleado = Empleado.query.get(id)
    if not empleado:
        flash("⚠ Empleado no encontrado.")
        return redirect(url_for("admin_panel"))

    if request.method == 'POST':
        nuevo_nombre = request.form['nombre'].strip().upper()
        nuevo_cedula = request.form['cedula'].strip()

        existente = Empleado.query.filter(Empleado.cedula == nuevo_cedula, Empleado.id != id).first()
        if existente:
            flash("⚠ Ya existe un empleado con esa cédula.")
            return redirect(url_for('editar_empleado', id=id))

        empleado.nombre = nuevo_nombre
        empleado.cedula = nuevo_cedula
        db.session.commit()

        flash("✔ Empleado actualizado correctamente.")
        return redirect(url_for('admin_panel'))

    return render_template("editar_empleado.html", empleado=empleado, config=cargar_config())

@app.route("/eliminar_empleado/<int:id>")
def eliminar_empleado(id):
    empleado = Empleado.query.get(id)
    if not empleado:
        flash("⚠ Empleado no encontrado.")
        return redirect(url_for("admin_panel"))

    db.session.delete(empleado)
    db.session.commit()
    flash("✔ Empleado eliminado correctamente.")
    return redirect(url_for("admin_panel"))


def obtener_registros_filtrados(cedula=None, fecha_inicio=None, fecha_fin=None):
    query = Registro.query

    if cedula:
        empleado = Empleado.query.filter_by(cedula=cedula.strip()).first()
        if empleado:
            query = query.filter(Registro.empleado == empleado.nombre)
        else:
            return []  # ← Cedula no existe

    if fecha_inicio:
        try:
            fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
            query = query.filter(Registro.fecha >= fi)
        except ValueError:
            pass

    if fecha_fin:
        try:
            ff = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
            query = query.filter(Registro.fecha <= ff)
        except ValueError:
            pass

    registros_raw = query.order_by(Registro.fecha, Registro.hora).all()
    registros_por_empleado_fecha = {}

    for r in registros_raw:
        clave = (r.cedula, r.empleado.upper(), r.fecha)
        if clave not in registros_por_empleado_fecha:
            registros_por_empleado_fecha[clave] = []
        registros_por_empleado_fecha[clave].append(r)

    registros_finales = []
    for (cedula, nombre, fecha), lista in registros_por_empleado_fecha.items():
        i = 0
        while i < len(lista):
            entrada = salida = None
            tiempo_total = estado = ""

            if lista[i].tipo == "entrada":
                entrada = lista[i].hora
                if i + 1 < len(lista) and lista[i + 1].tipo == "salida":
                    salida = lista[i + 1].hora
                    estado = "Entrada y Salida"
                    entrada_dt = datetime.combine(fecha, entrada)
                    salida_dt = datetime.combine(fecha, salida)
                    tiempo_total = str(salida_dt - entrada_dt).split('.')[0]
                    i += 2
                else:
                    estado = "Entrada"
                    i += 1
            elif lista[i].tipo == "salida":
                salida = lista[i].hora
                estado = "Salida"
                i += 1
            else:
                i += 1

            registros_finales.append({
                "cedula": cedula,
                "nombre": nombre,
                "fecha": fecha.strftime("%Y-%m-%d"),
                "hora_entrada": entrada.strftime("%H:%M:%S") if entrada else "",
                "hora_salida": salida.strftime("%H:%M:%S") if salida else "",
                "estado": estado,
                "tiempo_total": tiempo_total
            })

    return registros_finales

@app.route("/exportar_excel")
def exportar_excel():
    if session.get("rol") != "admin":
        return redirect(url_for("login_admin"))

    cedula = request.args.get("cedula", "").strip()
    fecha_inicio = request.args.get("fecha_inicio", "").strip()
    fecha_fin = request.args.get("fecha_fin", "").strip()

    registros = obtener_registros_filtrados(cedula, fecha_inicio, fecha_fin)

    if not registros:
        flash("⚠ No se encontraron registros con los filtros seleccionados.", "warning")
        return redirect(url_for("admin_panel"))

    df = pd.DataFrame(registros)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Historial")
    output.seek(0)

    return send_file(output, download_name="historial_registros.xlsx", as_attachment=True)

@app.route("/exportar_pdf")
def exportar_pdf():
    if session.get("rol") != "admin":
        return redirect(url_for("login_admin"))

    cedula = request.args.get("cedula", "").strip()
    fecha_inicio = request.args.get("fecha_inicio", "").strip()
    fecha_fin = request.args.get("fecha_fin", "").strip()

    registros = obtener_registros_filtrados(cedula, fecha_inicio, fecha_fin)

    if not registros:
        flash("⚠ No se encontraron registros con los filtros seleccionados.", "warning")
        return redirect(url_for("admin_panel"))

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Historial de Registro - Dragon Guard", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", "B", 10)

    encabezados = ["Cédula", "Nombre", "Fecha", "Hora Entrada", "Hora Salida", "Estado", "Tiempo"]
    anchos = [30, 65, 25, 25, 25, 35, 25]

    for i, encabezado in enumerate(encabezados):
        pdf.cell(anchos[i], 8, encabezado, 1, 0, "C")
    pdf.ln()

    pdf.set_font("Arial", "", 10)
    for r in registros:
        pdf.cell(anchos[0], 8, r["cedula"], 1)
        pdf.cell(anchos[1], 8, r["nombre"], 1)
        pdf.cell(anchos[2], 8, r["fecha"], 1)
        pdf.cell(anchos[3], 8, r["hora_entrada"], 1)
        pdf.cell(anchos[4], 8, r["hora_salida"], 1)
        pdf.cell(anchos[5], 8, r["estado"], 1)
        pdf.cell(anchos[6], 8, r["tiempo_total"], 1)
        pdf.ln()

    pdf_bytes = pdf.output(dest='S').encode('latin1')

    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={"Content-Disposition": "attachment; filename=historial.pdf"}
    )


@app.route("/registro_dashboard", methods=["POST"])
def registro_dashboard():
    cedula = request.form.get("cedula", "").strip()
    if not cedula:
        flash("⚠ Debe ingresar un número de cédula.", "warning")
        return redirect(url_for("index"))

    empleado = Empleado.query.filter_by(cedula=cedula).first()
    if not empleado:
        flash("⚠ FAVOR NOTIFICAR A SU ADMINISTRADOR DEL SISTEMA PARA CREAR AL EMPLEADO.", "danger")
        return redirect(url_for("index"))

    ahora = datetime.now()
    fecha = ahora.date()
    hora = ahora.time()

    ultimo = Registro.query.filter_by(cedula=cedula, fecha=fecha).order_by(Registro.hora.desc()).first()
    tipo = "entrada" if not ultimo or ultimo.tipo == "salida" else "salida"

    nuevo_registro = Registro(
        cedula=cedula,
        empleado=empleado.nombre,
        fecha=fecha,
        hora=hora,
        tipo=tipo
    )
    db.session.add(nuevo_registro)
    db.session.commit()

    flash(f"✔ {tipo.capitalize()} registrada exitosamente para {empleado.nombre}", "success")
    return redirect(url_for("index"))

def obtener_registros_procesados():
    registros_raw = Registro.query.order_by(Registro.fecha.desc(), Registro.hora.desc()).all()
    registros = []
    registros_por_empleado = {}

    for r in registros_raw:
        empleado_obj = Empleado.query.filter_by(nombre=r.empleado).first()
        cedula = empleado_obj.cedula if empleado_obj else '---'
        clave = (r.empleado, r.fecha)
        if clave not in registros_por_empleado:
            registros_por_empleado[clave] = []
        registros_por_empleado[clave].append(r)

    for (nombre, fecha), registros_list in registros_por_empleado.items():
        registros_list.sort(key=lambda x: x.hora)
        for i in range(len(registros_list)):
            r = registros_list[i]
            tiempo_total = "---"
            if r.tipo == "entrada":
                if i + 1 < len(registros_list) and registros_list[i + 1].tipo == "salida":
                    entrada = datetime.combine(r.fecha, r.hora)
                    salida = datetime.combine(registros_list[i + 1].fecha, registros_list[i + 1].hora)
                    delta = salida - entrada
                    tiempo_total = str(delta).split('.')[0]

            registros.append({
                "fecha": r.fecha,
                "hora": r.hora,
                "tipo": r.tipo,
                "nombre": r.empleado,
                "cedula": cedula,
                "tiempo_total": tiempo_total
            })

    return registros




@app.route("/guardar_auto_informe", methods=["POST"])
def guardar_auto_informe():
    activo = request.form.get("activo") == "true"
    hora = request.form.get("hora", "")

    config = {"activo": activo, "hora": hora}
    with open(AUTO_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4)

    if activo and hora:
        hora_dt = datetime.strptime(hora, "%H:%M").time()
        if scheduler.get_job("informe_diario"):
            scheduler.remove_job("informe_diario")

        scheduler.add_job(
            func=generar_informe_programado,
            trigger="cron",
            hour=hora_dt.hour,
            minute=hora_dt.minute,
            id="informe_diario",
            replace_existing=True
        )
    else:
        if scheduler.get_job("informe_diario"):
            scheduler.remove_job("informe_diario")

    return jsonify({"success": True})



@app.route("/auto_config")
def obtener_auto_config():
    if os.path.exists(AUTO_CONFIG_PATH):
        try:
            with open(AUTO_CONFIG_PATH, "r", encoding="utf-8") as f:
                return jsonify(json.load(f))
        except:
            pass
    return jsonify({"activo": False, "hora": ""})


# =====================
# CIERRE ORDENADO
# =====================

atexit.register(lambda: scheduler.shutdown())



# =====================
# EJECUCIÓN PRINCIPAL
# =====================

@app.route("/health")
def health():
    return "OK", 200


# 🔚 Bloque final que se ejecuta al iniciar la app (modo local o en Railway)
with app.app_context():
    crear_bd_si_no_existe()
    inicializar_auto_informe()

if __name__ == "__main__":
    print("🟢 Servidor Flask corriendo en modo desarrollo (local)")
    app.run(debug=True)
else:
    print("🟢 Servidor Flask iniciado por Gunicorn (producción)")


